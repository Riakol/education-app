import openpyxl.styles
import pandas as pd
import openpyxl
from database import requests
from decimal import Decimal, ROUND_UP


async def create_attendance_excel(data, eng_level, group_number, transfer_data, student_name_dict, start_date=None, end_date=None):
    attendance_dict = {}
    total_attendance = {}
    student_ids = set() 

    for row in data:
        name = row['student_name']
        student_ids.add(row['student_id'])
        day = int(row['day'])
        status = row['status']
        absence_reason = row.get('absence_reason', '') 
        date = pd.to_datetime(row['date'])  
        month_year = date.to_period('M')  

        if month_year not in attendance_dict:
            attendance_dict[month_year] = {}

        if name not in attendance_dict[month_year]:
            attendance_dict[month_year][name] = {}

        attendance_dict[month_year][name][day] = {
            'status': status,
            'absence_reason': absence_reason
        }

        if name not in total_attendance:
            total_attendance[name] = {'Present': 0, 'Absent': 0, 'Good Reason': 0}

        if status == 'present':
            total_attendance[name]['Present'] += 1
        elif status == 'absent':
            total_attendance[name]['Absent'] += 1
            if absence_reason: 
                total_attendance[name]['Good Reason'] += 1
        elif status == 'other':
            total_attendance[name]['Good Reason'] += 1

    if start_date and end_date:
        date = f"{start_date.strftime('%d.%m.%y')} - {end_date.strftime('%d.%m.%y')}"
    else:
        date = "Alltime"
    file_name = f"{eng_level}_{group_number}__{date}_attendance.xlsx"

    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        
        for month_year in sorted(attendance_dict.keys()):
            unique_days = sorted({day for days in attendance_dict[month_year].values() for day in days.keys()})

            attendance_rows = []

            for name, days in attendance_dict[month_year].items():
                attendance_row = {'Name': name}
                for day in unique_days:
                    day_info = days.get(day, {})
                    status = day_info.get('status', '')
                    absence_reason = day_info.get('absence_reason', '')

                    if status == 'present':
                        attendance_row[day] = '✅'  
                    elif status == 'absent':
                        attendance_row[day] = '❌' 
                    elif status == 'other':
                        attendance_row[day] = absence_reason if absence_reason else 'No reason'
                    else:
                        attendance_row[day] = ''
                
                attendance_rows.append(attendance_row)

            attendance_df = pd.DataFrame(attendance_rows)
            
            sheet_name = month_year.strftime('%B %Y')
      
            header_df = pd.DataFrame([[f"{eng_level}: {group_number}"]])
            header_df.to_excel(writer, index=False, header=False, sheet_name=sheet_name)
            attendance_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)

            worksheet = writer.sheets[sheet_name]
            worksheet.column_dimensions['A'].width = 20

            for col in range(2, 15):
                column_letter = openpyxl.utils.get_column_letter(col)
                worksheet.column_dimensions[column_letter].width = 5

            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

            for row in range(3, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 30

        total_rows = []
        max_total_days = 0
        # total_ielts_sum = Decimal('0.00')  # Для хранения общей суммы для IELTS
        total_subscription_sum = Decimal('0')

        for name, counts in total_attendance.items():
            total_days = counts['Present'] + counts['Absent']  
            subscription = Decimal('0')  

            match eng_level:
                case "IELTS":
                    student_id = next((row['student_id'] for row in data if row['student_name'] == name), None)
                    if student_id is not None:
                        payments = await requests.get_payment(student_id)
                        if payments:
                            amount = payments[0]['amount']
                            if total_days == 12:
                                subscription = amount
                            else:
                                subscription = (amount / Decimal('12')) * Decimal(total_days)
                                subscription = subscription.quantize(Decimal('1'), rounding=ROUND_UP)
                            total_subscription_sum += subscription
                        # print(f"Total IELTS: {total_subscription_sum}")
                case _:
                    if total_days > 0:
                        if total_days == 12:
                            total_subscription_sum = Decimal('70000')
                        else:
                            total_subscription_sum = (Decimal('70000') / Decimal('12')) * Decimal(max_total_days)
                            total_subscription_sum = total_subscription_sum.quantize(Decimal('1'), rounding=ROUND_UP)

                    max_total_days = max(max_total_days, total_days)

            # print(f"Subscription: {subscription}")
            subscription = subscription.quantize(Decimal('1'), rounding=ROUND_UP)
            total_row = {
                'Name': name,
                'Present': counts['Present'],
                'Absent': counts['Absent'],
                'Good Reason': counts['Good Reason'],
                'Salary': subscription  
            }
            total_rows.append(total_row)

        
        total_df = pd.DataFrame(total_rows)

        total_header_df = pd.DataFrame([[f"{eng_level}: {group_number}"]])
        total_header_df.to_excel(writer, index=False, header=False, sheet_name='Total')
        total_df.to_excel(writer, index=False, sheet_name='Total', startrow=2)

        worksheet = writer.sheets['Total']
        column_widths = {
            'A': 20,  
            'B': 15,  
            'C': 15,  
            'D': 20,
            'E': 15   
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        # Ограничиваем высоту строк в Total
        for row in range(3, worksheet.max_row + 1):  
            worksheet.row_dimensions[row].height = 30

        # Вставляем общую сумму подписок в последнюю строку колонки Subscription
        total_row_index = len(total_rows) + 4
        worksheet.cell(row=total_row_index, column=5, value=total_subscription_sum)  # Вставляем общую сумму

        total_cell = worksheet.cell(row=total_row_index, column=5)
        total_cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Желтый цвет

        # Ограничиваем высоту строки с общей суммой
        worksheet.row_dimensions[total_row_index].height = 30

        for row in range(4, len(total_rows) + 4):
            cell = worksheet.cell(row=row, column=5)  # Колонка E
            cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        transfer_rows = []
        transfer_present_count = {}
        transfer_good_reason_count = {}

        for transfer in transfer_data:
            student_name = student_name_dict.get(transfer['student_id'], 'Unknown')
            transfer_key = (student_name, transfer['eng_lvl'], transfer['group_name'])

            if transfer_key not in transfer_present_count:
                transfer_present_count[transfer_key] = 0
                transfer_good_reason_count[transfer_key] = 0 

            if transfer['status'] == 'present':
                transfer_present_count[transfer_key] += 1
            elif transfer['status'] == 'other' and transfer['absence_reason']:
                transfer_good_reason_count[transfer_key] += 1

        for (student_name, eng_lvl, group_number), present_count in transfer_present_count.items():
            good_reason_count = transfer_good_reason_count.get((student_name, eng_lvl, group_number), 0)
            transfer_row = {
                'Name': student_name,
                'Transfer': f"{eng_lvl}: {group_number}",  
                'Present': present_count,  
                'Good Reason': good_reason_count  
            }
            transfer_rows.append(transfer_row)

        if transfer_rows:
            transfer_df = pd.DataFrame(transfer_rows)
            transfer_df.to_excel(writer, index=False, sheet_name='Transfer')

            worksheet = writer.sheets['Transfer']
            column_widths = {
                'A': 20,  
                'B': 25,  
                'C': 15,  
                'D': 15   
            }

            for column, width in column_widths.items():
                worksheet.column_dimensions[column].width = width

            # Ограничиваем высоту строк в Transfer
            for row in range(2, worksheet.max_row + 1):  
                worksheet.row_dimensions[row].height = 30  

    return file_name




# async def create_attendance_excel(data, eng_level, group_number, transfer_data, student_name_dict, start_date=None, end_date=None):
#     attendance_dict = {}
#     total_attendance = {}

#     student_ids = set() 

#     for row in data:
#         name = row['student_name']
#         student_ids.add(row['student_id'])
#         day = int(row['day'])
#         status = row['status']
#         absence_reason = row.get('absence_reason', '') 
#         date = pd.to_datetime(row['date'])  

#         month_year = date.to_period('M')  

#         if month_year not in attendance_dict:
#             attendance_dict[month_year] = {}

#         if name not in attendance_dict[month_year]:
#             attendance_dict[month_year][name] = {}

#         attendance_dict[month_year][name][day] = {
#             'status': status,
#             'absence_reason': absence_reason
#         }

#         if name not in total_attendance:
#             total_attendance[name] = {'Present': 0, 'Absent': 0, 'Good Reason': 0}

#         if status == 'present':
#             total_attendance[name]['Present'] += 1
#         elif status == 'absent':
#             total_attendance[name]['Absent'] += 1
#             if absence_reason: 
#                 total_attendance[name]['Good Reason'] += 1
#         elif status == 'other':
#             total_attendance[name]['Good Reason'] += 1

#     if start_date and end_date:
#         date = f"{start_date.strftime('%d.%m.%y')} - {end_date.strftime('%d.%m.%y')}"
#     else:
#         date = "Alltime"
#     file_name = f"{eng_level}_{group_number}__{date}_attendance.xlsx"

#     with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        
#         for month_year in sorted(attendance_dict.keys()):
#             unique_days = sorted({day for days in attendance_dict[month_year].values() for day in days.keys()})

#             attendance_rows = []

#             for name, days in attendance_dict[month_year].items():
#                 attendance_row = {'Name': name}
#                 for day in unique_days:
#                     day_info = days.get(day, {})
#                     status = day_info.get('status', '')
#                     absence_reason = day_info.get('absence_reason', '')

#                     if status == 'present':
#                         attendance_row[day] = '✅'  
#                     elif status == 'absent':
#                         attendance_row[day] = '❌' 
#                     elif status == 'other':
#                         attendance_row[day] = absence_reason if absence_reason else 'No reason'
#                     else:
#                         attendance_row[day] = ''
                
#                 attendance_rows.append(attendance_row)

#             attendance_df = pd.DataFrame(attendance_rows)
            
#             sheet_name = month_year.strftime('%B %Y')
      
#             header_df = pd.DataFrame([[f"{eng_level}: {group_number}"]])
#             header_df.to_excel(writer, index=False, header=False, sheet_name=sheet_name)
#             attendance_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)

#             worksheet = writer.sheets[sheet_name]
#             worksheet.column_dimensions['A'].width = 20

#             for col in range(2, 15):
#                 column_letter = openpyxl.utils.get_column_letter(col)
#                 worksheet.column_dimensions[column_letter].width = 5

#             for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
#                 for cell in row:
#                     cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

#             for row in range(3, worksheet.max_row + 1):
#                 worksheet.row_dimensions[row].height = 30

#         total_rows = []
#         for name, counts in total_attendance.items():
#             total_row = {
#                 'Name': name,
#                 'Present': counts['Present'],
#                 'Absent': counts['Absent'],
#                 'Good Reason': counts['Good Reason']
#             }
#             total_rows.append(total_row)

#         total_df = pd.DataFrame(total_rows)

#         total_header_df = pd.DataFrame([[f"{eng_level}: {group_number}"]])
#         total_header_df.to_excel(writer, index=False, header=False, sheet_name='Total')
#         total_df.to_excel(writer, index=False, sheet_name='Total', startrow=2)

#         worksheet = writer.sheets['Total']
#         column_widths = {
#             'A': 20,  
#             'B': 15,  
#             'C': 15,  
#             'D': 20   
#         }

#         for column, width in column_widths.items():
#             worksheet.column_dimensions[column].width = width

#         # Ограничиваем высоту строк в Total
#         for row in range(3, worksheet.max_row + 1):  
#             worksheet.row_dimensions[row].height = 30

#         transfer_rows = []
#         transfer_present_count = {}
#         transfer_good_reason_count = {}

#         for transfer in transfer_data:
#             student_name = student_name_dict.get(transfer['student_id'], 'Unknown')
#             transfer_key = (student_name, transfer['eng_lvl'], transfer['group_name'])

#             if transfer_key not in transfer_present_count:
#                 transfer_present_count[transfer_key] = 0
#                 transfer_good_reason_count[transfer_key] = 0 

#             if transfer['status'] == 'present':
#                 transfer_present_count[transfer_key] += 1
#             elif transfer['status'] == 'other' and transfer['absence_reason']:
#                 transfer_good_reason_count[transfer_key] += 1

#         for (student_name, eng_lvl, group_number), present_count in transfer_present_count.items():
#             good_reason_count = transfer_good_reason_count.get((student_name, eng_lvl, group_number), 0)
#             transfer_row = {
#                 'Name': student_name,
#                 'Transfer': f"{eng_lvl}: {group_number}",  
#                 'Present': present_count,  
#                 'Good Reason': good_reason_count  
#             }
#             transfer_rows.append(transfer_row)

#         if transfer_rows:
#             transfer_df = pd.DataFrame(transfer_rows)

#             transfer_df.to_excel(writer, index=False, sheet_name='Transfer')

#             worksheet = writer.sheets['Transfer']
#             column_widths = {
#                 'A': 20,  
#                 'B': 25,  
#                 'C': 15,  
#                 'D': 15   
#             }

#             for column, width in column_widths.items():
#                 worksheet.column_dimensions[column].width = width

#             # Ограничиваем высоту строк в Transfer
#             for row in range(2, worksheet.max_row + 1):  
#                 worksheet.row_dimensions[row].height = 30  

#     return file_name
